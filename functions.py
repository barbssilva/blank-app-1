import os
import openpyxl
import copy
from openpyxl.utils import get_column_letter
import xlwings as xw


#função para juntar os excels
def join_excels(arquivos,tipo_pl, output_file):
    # Abrir primeiro arquivo para template e header
    template_wb = openpyxl.load_workbook(f'{arquivos[0]}')
    template_ws = template_wb.active

    # Criar novo workbook e copiar formatação do template
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    # Copiar as primeiras 12 linhas com formatação
    for row in range(1, 13):
        for col in range(1, template_ws.max_column + 1):
            source_cell = template_ws.cell(row=row, column=col)
            target_cell = new_ws.cell(row=row, column=col)
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell.font = copy.copy(source_cell.font)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.number_format = source_cell.number_format
                target_cell.alignment = copy.copy(source_cell.alignment)

    # Linha atual para adicionar dados
    current_row = 13

    # Processar todos os arquivos
    for file in arquivos:
        wb = openpyxl.load_workbook(f'{file}')
        ws = wb.active
        
        # Começar da linha 13 de cada arquivo
        for row in range(13, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                source_cell = ws.cell(row=row, column=col)
                target_cell = new_ws.cell(row=current_row, column=col)
                target_cell.value = source_cell.value
                if source_cell.has_style:
                    target_cell.font = copy.copy(source_cell.font)
                    target_cell.border = copy.copy(source_cell.border)
                    target_cell.fill = copy.copy(source_cell.fill)
                    target_cell.number_format = source_cell.number_format
                    target_cell.alignment = copy.copy(source_cell.alignment)
            current_row += 1

    # Ajustar largura das colunas
    for col in range(1, template_ws.max_column + 1):
        new_ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = \
            template_ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width

    # Remover linhas totalmente vazias (apenas a partir da linha 13)
    removed = 0
    for row_idx in range(new_ws.max_row, 12, -1):  # iterar de baixo para cima
        is_blank = True
        for cell in new_ws[row_idx]:
            v = cell.value
            if v is not None and str(v).strip() != "":
                is_blank = False
                break
        if is_blank:
            new_ws.delete_rows(row_idx, 1)
            removed += 1

    if tipo_pl == 'standard':
        # Somar inteiros da coluna C nas linhas onde a coluna A é "NUMBER OF BOXES:",
        # manter apenas a última ocorrência e escrever a soma na coluna C dessa linha.
        rows_with_label = []
        total = 0

        for r in range(13, new_ws.max_row + 1):
            a = new_ws.cell(row=r, column=1).value
            if a is None:
                continue
            if str(a).strip().upper() == "NUMBER OF BOXES:":
                rows_with_label.append(r)
                c_val = new_ws.cell(row=r, column=3).value
                n = int(c_val) if c_val not in (None, "") else 0
                total += n

        if rows_with_label:
            last_row = max(rows_with_label)
            to_delete = [r for r in rows_with_label if r != last_row]
            to_delete.sort(reverse=True)
            for r in to_delete:
                new_ws.delete_rows(r, 1)
            # calcular nova posição da última linha após as deleções anteriores
            deleted_before_last = sum(1 for d in to_delete if d < last_row)
            new_last_row = last_row - deleted_before_last

            # escrever soma (inteiro) na coluna C da última linha, mantendo estilos
            tgt = new_ws.cell(row=new_last_row, column=3)
            tgt.value = int(total)

            # colocar fórmula na coluna L (coluna 12): soma de L13 até à linha anterior ao "NUMBER OF BOXES:"
            for col in range(12, 32):  # 12 = L, 31 = AE
                if col == 13:  # pular M
                    continue
                col_letter = get_column_letter(col)
                formula_cell = new_ws.cell(row=new_last_row, column=col)
                if new_last_row > 13:
                    formula_cell.value = f"=SUM({col_letter}13:{col_letter}{new_last_row-1})"
                else:
                    formula_cell.value = 0
                
                formula_cell.number_format = '0'
        
        new_ws.column_dimensions['D'].width = 30

        new_ws.title = 'Standard PL'


        # Salvar arquivo final
        #output_file = 'packing_lists/STANDARD_PL.xlsx'

    elif tipo_pl == 'summary':
        #remover linhas com "TOTAL" e manter apenas a última ocorrência
        rows_with_label = []
        total = 0

        for r in range(13, new_ws.max_row + 1):
            a = new_ws.cell(row=r, column=5).value
            if a is None:
                continue
            if str(a).strip().upper() == "TOTAL":
                rows_with_label.append(r)

        if rows_with_label:
            last_row = max(rows_with_label)
            to_delete = [r for r in rows_with_label if r != last_row]
            to_delete.sort(reverse=True)
            for r in to_delete:
                new_ws.delete_rows(r, 1)
            # calcular nova posição da última linha após as deleções anteriores
            deleted_before_last = sum(1 for d in to_delete if d < last_row)
            new_last_row = last_row - deleted_before_last

            # colocar fórmula na coluna j (coluna 8): soma de L13 até à linha anterior ao "TOTAL"
            for col in range(8, 30):  # 8 = J, 29 = AC
                if col == 9:  # pular K
                    continue
                col_letter = get_column_letter(col)
                formula_cell = new_ws.cell(row=new_last_row, column=col)
                if new_last_row > 9:
                    formula_cell.value = f"=SUM({col_letter}13:{col_letter}{new_last_row-1})"
                else:
                    formula_cell.value = 0
                    
                formula_cell.number_format = '0'

        new_ws.column_dimensions['C'].width = 30
        new_ws.title = 'Summary PL'
        #output_file = 'packing_lists/SUMMARY_PL.xlsx'

    # Salvar arquivo final
    new_wb.save(output_file)
        
    return output_file

#função que junta as duas PLS - standard e summary
def join_pls(standard_pl, summary_pl,last_file):
    app = xw.App(visible=False)  # invisível
    wb_final = xw.Book()

    for ficheiro in [standard_pl, summary_pl]:
        wb = xw.Book(ficheiro)
        for sheet in wb.sheets:
            sheet.api.Copy(Before=wb_final.sheets[0].api)
        wb.close()

    try:
        wb_final.sheets["Folha1"].delete()
    except:
        pass

    # guardar
    wb_final.save(last_file)
    wb_final.close()
    app.quit()
    
    return

#remover ficheiros temporários
def remove_pls(standard_pl,summary_pl):
    os.remove(standard_pl)
    os.remove(summary_pl)

    return
