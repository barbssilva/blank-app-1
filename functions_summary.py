import os
import openpyxl
import copy
from openpyxl.utils import get_column_letter
import re
import pandas as pd
import streamlit as st


### SUMMARY PACKING LIST ###
#pre processamento dos excels summary packing list

def pre_proc_summary(arquivos):
    for file in arquivos:
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        max_row = ws.max_row
       # Mapeamento dos tamanhos para a coluna correta
        mapa_colunas = {
                '3m': 10,  # J
                '2y':10,  # J
                '3-6':10,  # J
                '6m':11,  # K
                '4':11,  # K
                '4y':11,  # K
                '6-9':11,  # K
                '9m':12,  # L
                '9-12':12,  # L
                '12m':13, # M
                '6':13, # M
                '6y':13, # M
                '12-18':13, # M
                '18m':14, # N
                '18-24':14, # N
                '24m':15, # O
                '24-36':15, # O
                '8':15, # O
                '8y':15, # O
                '36m':16, # P
                '10':17, # Q
                '10y':17, # Q
                '12':19, # S
                '12y':19, # S                
                '14':21, # U
                '14y':21, # U
                '16':23, # W
                '16y':23, # W
            }
        
        dados_movidos = []

        # Guardar dados das colunas J até X num dataframe antes de limpar
        for col in range(10, 25):  # J até X
            cabecalho = ws.cell(row=12, column=col).value

            if cabecalho is None or str(cabecalho).strip() == '':
                continue

            cabecalho_norm = str(cabecalho).strip().lower().replace(' ', '')

            if cabecalho_norm not in mapa_colunas:
                continue

            coluna_destino = mapa_colunas[cabecalho_norm]

            for row in range(13, max_row):
                valor = ws.cell(row=row, column=col).value

                # parar quando encontrar a primeira célula vazia nessa coluna
                if valor is None or str(valor).strip() == '':
                    continue

                dados_movidos.append({
                    'row': row,
                    'source_col': col,
                    'source_header': cabecalho,
                    'target_col': coluna_destino,
                    'value': valor,
                })

        df_movimentos = pd.DataFrame(dados_movidos)

        # Limpar colunas J até X, da linha 13 para baixo
        for row in range(13, max_row):
            for col in range(10, 25):  # J até X
                ws.cell(row=row, column=col).value = ''

        # Recolocar os dados nos locais certos
        for _, item in df_movimentos.iterrows():
            ws.cell(
                row=int(item['row']),
                column=int(item['target_col'])
            ).value = item['value']
        

        # Limpar linha 12, colunas J até X
        for col in range(10, 25):
            ws.cell(row=12, column=col).value = ''
        for row in range(13, max_row + 1):
            #selecionar a coluna w.o number e separar em 3 valores, w.o ente, w.o year, wo.number
            cell = ws.cell(row=row, column=3) #coluna w.o number
            val = cell.value
            if not isinstance(val, str):
                continue
                
            val = str(val).strip()    
            # Verifica se contém CH e BM em qualquer posição, maiúsculas/minúsculas
            m = re.search(r'CH[^0-9]*BM', val, flags=re.IGNORECASE)

            if m:
                ws.cell(row=row, column=1).value = "CH_BM"
                # Remove tudo o que não for número
                numeros = re.sub(r'[^0-9]', '', val)
                if len(numeros) >= 7:
                    wo_year = numeros[:4] #primeiros 4 numeros para o ano
                    wo_number = numeros[4:] #restantes numeros

                    ws.cell(row=row, column=2).value = wo_year
                    ws.cell(row=row, column=3).value = wo_number
            else:
                st.write("ANTES DE FAZER O PL STANDARD POR FAVOR CORRIJA QUALQUER ERRO NO PL SUMMARY!!!")
                st.write(f"ERRO!!{val} não está no formato esperado")
                    
            #selecionar a coluna style name e separar em 2 valores, style name e article number
            cell2 = ws.cell(row=row, column=5) #coluna STYLE NAME
            val2 = cell2.value
            if not isinstance(val2, str):
                continue

            m2 = re.search(r'ART', val2, flags=re.IGNORECASE)
            if m2:
                start = m2.start()
                end = m2.end() # retorna a posição inicial da correspondência
                first = val2[:start].strip()
                ws.cell(row=row, column=5).value = first

                second = val2[end:].strip().replace('.','')
                ws.cell(row=row, column=6).value  = second 
            else:
                st.write(f"ERRO!! O STYLE NAME {val2} não menciona o valor ART. A coluna ARTICLE ficou vazia") 
                ws.cell(row=row, column=6).value  = ''

        wb.save(file)
    return

#função para juntar os excels
def join_excels(arquivos, output_file):
    # Abrir primeiro arquivo para template e header
    template_wb = openpyxl.load_workbook(arquivos[0])
    template_ws = template_wb.active

    # Criar novo workbook e copiar formatação do template
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    # Copiar as primeiras 12 linhas com formatação
    for row in range(1, 13):
        for col in range(1, template_ws.max_column + 1):
            source_cell = template_ws.cell(row=row, column=col)
            target_cell = new_ws.cell(row=row, column=col)
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell.font = copy.copy(source_cell.font)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.number_format = source_cell.number_format
                target_cell.alignment = copy.copy(source_cell.alignment)

    # Linha atual para adicionar dados
    current_row = 13

    # Processar todos os arquivos
    for file in arquivos:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # Começar da linha 13 de cada arquivo
        for row in range(13, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                source_cell = ws.cell(row=row, column=col)
                target_cell = new_ws.cell(row=current_row, column=col)
                target_cell.value = source_cell.value
                if source_cell.has_style:
                    target_cell.font = copy.copy(source_cell.font)
                    target_cell.border = copy.copy(source_cell.border)
                    target_cell.fill = copy.copy(source_cell.fill)
                    target_cell.number_format = source_cell.number_format
                    target_cell.alignment = copy.copy(source_cell.alignment)
            current_row += 1

    # Ajustar largura das colunas
    for col in range(1, template_ws.max_column + 1):
        new_ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = \
            template_ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width

    # Remover linhas totalmente vazias (apenas a partir da linha 13)
    removed = 0
    for row_idx in range(new_ws.max_row, 12, -1):  # iterar de baixo para cima
        is_blank = True
        for cell in new_ws[row_idx]:
            v = cell.value
            if v is not None and str(v).strip() != "":
                is_blank = False
                break
        if is_blank:
            new_ws.delete_rows(row_idx, 1)
            removed += 1

    #remover linhas com "TOTAL" e manter apenas a última ocorrência
    rows_with_label = []
    #total = 0
        #print(new_ws.max_row)
    for r in range(13, new_ws.max_row + 1):
        a = new_ws.cell(row=r, column=7).value
        if a is None:
            continue
        if str(a).strip().upper() == "TOTAL":
            rows_with_label.append(r)

    if rows_with_label:
        last_row = max(rows_with_label)
        to_delete = [r for r in rows_with_label if r != last_row]
        to_delete.sort(reverse=True)
        for r in to_delete:
            new_ws.delete_rows(r, 1)
        
    new_ws.column_dimensions['E'].width = 30

    new_ws.title = 'Summary PL'
    
    # Salvar arquivo final
    new_wb.save(output_file)
        
    return output_file


def ordenar_summary(final_file):
    wb = openpyxl.load_workbook(final_file)
    ws = wb.active

    wo_col=3  #coluna W.O NUMBER
    start_row = 13
    data_rows = []
    for row in range(start_row, ws.max_row):
        row_values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        if set(row_values) == {None}:
            break
        data_rows.append(row_values)


    data_rows.sort(key=lambda r: (r[wo_col - 1] or ""))

    for i, row_values in enumerate(data_rows, start=start_row):
        for col, value in enumerate(row_values, start=1):
            ws.cell(row=i, column=col).value = value

    n_rows = len(data_rows)
    unique_rows = {}


    #ideia tentar encontrar onde é o ultimo ch_bm na primeira coluna e essa será a ultima linha a considerar
    # Percorrer linhas a partir da linha 13
    for row in ws.iter_rows(min_row=13, max_row=13+n_rows-1, values_only=True):
        key = row[0:7]  # colunas 1 a 7

        if key in unique_rows:
            # somar colunas 8 e 10 a 16
            for col_idx in [7] + list(range(9, 16)):
                val = row[col_idx] or 0
                unique_rows[key][col_idx] += val
        else:
            # criar cópia mutável da linha
            new_row = list(row)
            for col_idx in [7] + list(range(9, 16)):
                new_row[col_idx] = new_row[col_idx] or 0
            unique_rows[key] = new_row
                
    # Escrever as linhas únicas de volta, mas antes colocar todas as linhas limpas
    for clear_row in range(13, 13 + n_rows):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=clear_row, column=col).value = None
            
    row_cursor = 13
    for row in unique_rows.values():
        for col_idx, value in enumerate(row):
            ws.cell(row=row_cursor, column=col_idx + 1, value=value)
        row_cursor += 1
    
    last_row = ws.max_row
    for col in range(8, 25):  # 12 = L, 31 = AE
        if col == 9:  # pular M
            continue
        col_letter = get_column_letter(col)
        formula_cell = ws.cell(row=last_row, column=col)
        if last_row > 13:
            formula_cell.value = f"=SUM({col_letter}13:{col_letter}{last_row-1})"
        else:
            formula_cell.value = 0
                
        formula_cell.number_format = '0'
    wb.save(final_file)
    return
