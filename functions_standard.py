import os
import openpyxl
import copy
from openpyxl.utils import get_column_letter
import re
import pandas as pd
import streamlit as st

### Standard PACKING LIST ###
#pre processamento dos excels standard packing list

def pre_proc_standard(arquivos):
    for file in arquivos:
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        max_row = ws.max_row
        # Mapeamento dos tamanhos para a coluna correta
        mapa_colunas = {
                '3m': 12,  # L
                '2y':12,  # L
                '3-6':12,  # L
                '6m':13,  # M
                '4':13,  # M
                '4y':13,  # M
                '6-9':13,  # M
                '9m':14,  # N
                '9-12':14,  # N
                '12m':15, # O
                '6':15, # O
                '6y':15, # O
                '12-18':15, # O
                '18m':16, # P
                '18-24':16, # P
                '24m':17, # Q
                '24-36':17, # Q
                '8':17, # Q
                '8y':17, # Q
                '36m':18, # R
                '10':19, # S
                '10y':19, # S
                '12':21, # U
                '12y':21, # U                
                '14':23, # W
                '14y':23, # W
                '16':25, # Y
                '16y':25, # Y
            }
        
        dados_movidos = []

        # Guardar dados das colunas K até Z num dataframe antes de limpar
        for col in range(11, 26):  # K até Z
            cabecalho = ws.cell(row=12, column=col).value

            if cabecalho is None or str(cabecalho).strip() == '':
                continue

            cabecalho_norm = str(cabecalho).strip().lower().replace(' ', '')

            if cabecalho_norm not in mapa_colunas:
                continue

            coluna_destino = mapa_colunas[cabecalho_norm]

            for row in range(13, max_row):
                valor = ws.cell(row=row, column=col).value

                # parar quando encontrar a primeira célula vazia nessa coluna
                if valor is None or str(valor).strip() == '':
                    continue

                dados_movidos.append({
                    'row': row,
                    'source_col': col,
                    'source_header': cabecalho,
                    'target_col': coluna_destino,
                    'value': valor,
                })

        df_movimentos = pd.DataFrame(dados_movidos)

        # Limpar colunas K até Z, da linha 13 para baixo
        for row in range(13, max_row):
            for col in range(11, 26):  # K até Z
                ws.cell(row=row, column=col).value = ''

        # Recolocar os dados nos locais certos
        for _, item in df_movimentos.iterrows():
            ws.cell(
                row=int(item['row']),
                column=int(item['target_col'])
            ).value = item['value']
        

        # Limpar linha 12, colunas K até Z
        for col in range(11, 26):
            ws.cell(row=12, column=col).value = ''
        for row in range(13, max_row + 1):
            #selecionar a coluna style name e separar em 2 valores, style name e article
            cell2 = ws.cell(row=row, column=3) #coluna STYLE NAME
            val2 = cell2.value
            if not isinstance(val2, str):
                continue

            m2 = re.search(r'ART', val2, flags=re.IGNORECASE)
            if m2:
                start = m2.start()
                end = m2.end() # retorna a posição inicial da correspondência
                first = val2[:start].strip()
                ws.cell(row=row, column=3).value = first

                second = val2[end:].strip().replace('.','')
                ws.cell(row=row, column=4).value  = second 
            else:
                st.write(f'ERRO!! O STYLE NAME {val2} não menciona o valor ART. A coluna ARTICLE ficou vazia') 
                ws.cell(row=row, column=4).value  = ''

        wb.save(file)
    return


def join_excels_standard(arquivos,output_file):
    # Abrir primeiro arquivo para template e header
    template_wb = openpyxl.load_workbook(arquivos[0])
    template_ws = template_wb.active

    # Criar novo workbook e copiar formatação do template
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    # Copiar as primeiras 12 linhas com formatação
    for row in range(1, 13):
        for col in range(1, template_ws.max_column + 1):
            source_cell = template_ws.cell(row=row, column=col)
            target_cell = new_ws.cell(row=row, column=col)
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell.font = copy.copy(source_cell.font)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.number_format = source_cell.number_format
                target_cell.alignment = copy.copy(source_cell.alignment)

    # Linha atual para adicionar dados
    current_row = 13

    # Processar todos os arquivos
    for file in arquivos:
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        #print(ws.max_row)
        
        # Começar da linha 13 de cada arquivo
        for row in range(13, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                source_cell = ws.cell(row=row, column=col)
                target_cell = new_ws.cell(row=current_row, column=col)
                target_cell.value = source_cell.value
                if source_cell.has_style:
                    target_cell.font = copy.copy(source_cell.font)
                    target_cell.border = copy.copy(source_cell.border)
                    target_cell.fill = copy.copy(source_cell.fill)
                    target_cell.number_format = source_cell.number_format
                    target_cell.alignment = copy.copy(source_cell.alignment)
            current_row += 1

    # Ajustar largura das colunas
    for col in range(1, template_ws.max_column + 1):
        new_ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = \
            template_ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width

    # Remover linhas totalmente vazias (apenas a partir da linha 13)
    removed = 0
    #print(new_ws.max_row)
    for row_idx in range(new_ws.max_row, 12, -1):  # iterar de baixo para cima
        is_blank = True
        for cell in new_ws[row_idx]:
            v = cell.value
            if v is not None and str(v).strip() != "":
                is_blank = False
                break
        if is_blank:
            new_ws.delete_rows(row_idx, 1)
            removed += 1

    # Somar inteiros da coluna C nas linhas onde a coluna A é "NUMBER OF BOXES:",
    # manter apenas a última ocorrência e escrever a soma na coluna C dessa linha.
    rows_with_label = []
    total = 0

    for r in range(13, new_ws.max_row + 1):
        a = new_ws.cell(row=r, column=1).value
        if a is None:
            continue
        if str(a).strip().upper() == "NUMBER OF BOXES:":
            rows_with_label.append(r)
            c_val = new_ws.cell(row=r, column=3).value
            n = int(c_val) if c_val not in (None, "") else 0
            total += n

    if rows_with_label:
        last_row = max(rows_with_label)
        to_delete = [r for r in rows_with_label if r != last_row]
        to_delete.sort(reverse=True)
        for r in to_delete:
            new_ws.delete_rows(r, 1)
            # calcular nova posição da última linha após as deleções anteriores
            #deleted_before_last = sum(1 for d in to_delete if d < last_row)
        new_last_row = last_row - len(to_delete)

            # escrever soma (inteiro) na coluna C da última linha, mantendo estilos
        tgt = new_ws.cell(row=new_last_row, column=3)
        tgt.value = int(total)
        
    new_ws.column_dimensions['D'].width = 30

    new_ws.title = 'Standard PL'

    # Salvar arquivo final
    new_wb.save(output_file)
        
    return output_file

def ordenar_standard(path_summary, path_standard):
    #ordenar standard usando summary
    wb_ordem = openpyxl.load_workbook(path_summary, data_only=True)
    ws_ordem = wb_ordem.active

    ordem = []
    for row in ws_ordem.iter_rows(min_row=13, max_row=ws_ordem.max_row-1):  # ajusta se não tiver header
        if [row[4].value, row[6].value] in ordem:
            continue
        else:
            ordem.append([row[4].value, row[6].value])  # coluna 3 e 4
    
    wb = openpyxl.load_workbook(path_standard,data_only=True)
    ws = wb.active

    linhas = []

    for row in ws.iter_rows(min_row=13, max_row=ws.max_row-1):  # ajusta se não tiver header
        #chave = row[2].value  # coluna 3
        linhas.append(row)

    linhas_ordenadas = []
    for ordem_item in ordem:
        for row in linhas:
            if row[2].value == ordem_item[0] and row[4].value == ordem_item[1]:
                # Guardar os valores das células, não as referências
                linhas_ordenadas.append([cell.value for cell in row])
    
    # Escrever as linhas únicas de volta
    row_cursor = 13
    for row_n, row in enumerate(linhas_ordenadas):
        for col_idx, cell in enumerate(row):
            #primeiro colocar o valor de cada celula como string vazia, para não correr o risco de manter valores antigos dessa celula
            celula=ws.cell(row=row_cursor, column=col_idx+1, value="")
            celula=ws.cell(row=row_cursor, column=col_idx+1, value=cell)
            celula.alignment = Alignment(horizontal='center', vertical='center')
        row_cursor += 1
        
    # colocar fórmula na coluna L (coluna 12): soma de L13 até à linha anterior ao "NUMBER OF BOXES:"
    last_row = ws.max_row
    for col in range(7, 27):  # G ate Z
        if col == 10 or col == 11:  # pular j e k
            continue
        col_letter = get_column_letter(col)
        formula_cell = ws.cell(row=last_row, column=col)
        if last_row > 13:
            formula_cell.value = f"=SUM({col_letter}13:{col_letter}{last_row-1})"
        else:
            formula_cell.value = 0
        if col == 7 or col == 8:
            formula_cell.number_format = '0.00'    
            # centrar horizontal e verticalmente
            formula_cell.alignment = Alignment(
            horizontal='center',
            vertical='center')
        else:
            formula_cell.number_format = '0'

    wb.save(path_standard)
    
    return
    
